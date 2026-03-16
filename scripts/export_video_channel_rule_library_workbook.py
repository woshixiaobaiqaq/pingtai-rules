from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/m/亿一的AI小助理/open-codex/platform-content-audit")
RULES_PATH = ROOT / "data/rule_library/video_channel/rules.json"
CATALOG_PATH = ROOT / "data/rule_library/video_channel/catalog.json"
MANIFEST_PATH = ROOT / "data/rule_library/manifest.json"
OUTPUT_DIR = ROOT / "output/spreadsheet"
OUTPUT_PATH = OUTPUT_DIR / "video_channel_rule_library_complete_2026-03-16.xlsx"

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(value.replace("\n", " ").split())


def join_values(values: list[Any] | tuple[Any, ...] | None) -> str:
    if not values:
        return ""
    return " / ".join(str(value) for value in values if value)


def truncate(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def derive_source_group(rule: dict[str, Any]) -> str:
    source_type = str(rule.get("metadata", {}).get("source_type", ""))
    if source_type == "official_operation_standards":
        return "微信视频号运营规范"
    return "视频号常见违规内容概览"


def derive_primary_category(rule: dict[str, Any]) -> str:
    metadata = rule.get("metadata", {})
    if metadata.get("section_title"):
        return str(metadata["section_title"])
    title = str(rule.get("title", ""))
    if "：" in title:
        return title.split("：", 1)[0]
    return "未分类"


def derive_topic(rule: dict[str, Any]) -> str:
    title = str(rule.get("title", ""))
    if "：" in title:
        return title.split("：", 1)[1]
    return title


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_table_style(sheet, wrap_columns: set[int]) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in wrap_columns)


def set_widths(sheet, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def build_overview_sheet(workbook: Workbook, rules: list[dict[str, Any]], catalog: dict[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "库总览"
    sheet.append(["指标", "数值"])

    severity_counter = Counter(str(rule.get("severity", "")) for rule in rules)
    source_counter = Counter(derive_source_group(rule) for rule in rules)
    image_rule_count = sum(1 for rule in rules if rule.get("metadata", {}).get("image_assets"))

    rows = [
        ("生成时间", datetime.now().isoformat(timespec="seconds")),
        ("规则总数", len(rules)),
        ("视频号常见违规内容概览规则", source_counter["视频号常见违规内容概览"]),
        ("微信视频号运营规范规则", source_counter["微信视频号运营规范"]),
        ("高风险规则", severity_counter["high"]),
        ("中风险规则", severity_counter["medium"]),
        ("低风险规则", severity_counter["low"]),
        ("带本地图片规则数", image_rule_count),
        ("本地图片总数", catalog.get("image_count", 0)),
        ("来源页数量", len(catalog.get("source_notes", []))),
        ("规则文件", str(RULES_PATH)),
        ("目录文件", str(CATALOG_PATH)),
    ]
    for row in rows:
        sheet.append(list(row))

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={2})
    set_widths(sheet, {1: 28, 2: 120})


def build_rule_table_sheet(workbook: Workbook, rules: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("完整规则总表")
    sheet.append(
        [
            "序号",
            "规则ID",
            "来源分组",
            "一级分类",
            "条款编号",
            "标题",
            "风险等级",
            "规则正文",
            "关键词",
            "正则",
            "标签",
            "推断候选标签",
            "主来源",
            "图片数",
            "视频案例数",
            "引用链接数",
            "图片本地路径",
        ]
    )

    for index, rule in enumerate(rules, start=1):
        metadata = rule.get("metadata", {})
        image_assets = metadata.get("image_assets") or []
        video_case_links = metadata.get("video_case_links") or []
        reference_links = metadata.get("reference_links") or []
        sheet.append(
            [
                index,
                rule.get("rule_id", ""),
                derive_source_group(rule),
                derive_primary_category(rule),
                metadata.get("official_rule_code", ""),
                rule.get("title", ""),
                rule.get("severity", ""),
                normalize_text(rule.get("content")),
                join_values(rule.get("keywords")),
                join_values(rule.get("regex_patterns")),
                join_values(rule.get("tags")),
                join_values(metadata.get("inferred_candidate_tags")),
                rule.get("source_url") or "",
                len(image_assets),
                len(video_case_links),
                len(reference_links),
                join_values([asset.get("path", "") for asset in image_assets]),
            ]
        )

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={6, 8, 9, 10, 11, 12, 13, 17})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(
        sheet,
        {
            1: 8,
            2: 16,
            3: 18,
            4: 20,
            5: 12,
            6: 34,
            7: 10,
            8: 92,
            9: 26,
            10: 26,
            11: 26,
            12: 26,
            13: 44,
            14: 10,
            15: 10,
            16: 10,
            17: 72,
        },
    )


def build_category_summary_sheet(workbook: Workbook, rules: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("分类汇总")
    sheet.append(
        [
            "来源分组",
            "一级分类",
            "规则数",
            "高风险",
            "中风险",
            "低风险",
            "示例条目",
        ]
    )

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for rule in rules:
        grouped[(derive_source_group(rule), derive_primary_category(rule))].append(rule)

    for (source_group, category), items in sorted(grouped.items()):
        severity_counter = Counter(str(item.get("severity", "")) for item in items)
        sample_titles = " / ".join(derive_topic(item) for item in items[:6])
        sheet.append(
            [
                source_group,
                category,
                len(items),
                severity_counter["high"],
                severity_counter["medium"],
                severity_counter["low"],
                sample_titles,
            ]
        )

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={7})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(sheet, {1: 20, 2: 22, 3: 10, 4: 10, 5: 10, 6: 10, 7: 88})


def build_image_assets_sheet(workbook: Workbook, rules: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("图片资产")
    sheet.append(
        [
            "规则ID",
            "来源分组",
            "一级分类",
            "标题",
            "图片说明",
            "本地路径",
            "原始链接",
        ]
    )

    for rule in rules:
        image_assets = rule.get("metadata", {}).get("image_assets") or []
        for asset in image_assets:
            sheet.append(
                [
                    rule.get("rule_id", ""),
                    derive_source_group(rule),
                    derive_primary_category(rule),
                    derive_topic(rule),
                    asset.get("caption", ""),
                    asset.get("path", ""),
                    asset.get("remote_url", ""),
                ]
            )

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={4, 5, 6, 7})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(sheet, {1: 16, 2: 20, 3: 18, 4: 28, 5: 28, 6: 66, 7: 66})


def build_source_sheet(workbook: Workbook, manifest: dict[str, Any], catalog: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("来源说明")
    sheet.append(["项目", "内容"])
    sheet.append(["规则清单", str(RULES_PATH)])
    sheet.append(["目录清单", str(CATALOG_PATH)])
    sheet.append(["规则库版本", str(manifest.get("version", ""))])
    sheet.append(["平台描述", str(manifest.get("description", ""))])
    sheet.append(["目录标题", str(catalog.get("title", ""))])
    sheet.append(["目录范围", str(catalog.get("scope", ""))])

    for index, source in enumerate(catalog.get("source_notes", []), start=1):
        prefix = f"来源 {index}"
        sheet.append([f"{prefix} 标题", str(source.get("title", ""))])
        sheet.append([f"{prefix} 链接", str(source.get("url", ""))])
        sheet.append([f"{prefix} 本地归档", str(source.get("local_archive", ""))])

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={2})
    set_widths(sheet, {1: 24, 2: 126})


def main() -> None:
    rules = read_json(RULES_PATH)["rules"]
    catalog = read_json(CATALOG_PATH)
    manifest = read_json(MANIFEST_PATH)

    workbook = Workbook()
    build_overview_sheet(workbook, rules, catalog)
    build_rule_table_sheet(workbook, rules)
    build_category_summary_sheet(workbook, rules)
    build_image_assets_sheet(workbook, rules)
    build_source_sheet(workbook, manifest, catalog)

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(str(OUTPUT_PATH))


if __name__ == "__main__":
    main()
