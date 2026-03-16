from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/m/亿一的AI小助理/open-codex/platform-content-audit")
MERGED_DIR = (
    ROOT
    / "data/source_archives/xiaohongshu_rule_baike"
    / "sequential_content_merged_20260312_173210"
)
NOTES_INDEX_PATH = MERGED_DIR / "notes_index.json"
OCR_DIR = MERGED_DIR / "notes_ocr"
OUTPUT_PATH = (
    ROOT
    / "output/spreadsheet"
    / "xiaohongshu_rule_baike_notes_2026-03-12.xlsx"
)


CATEGORY_INFO = {
    "内容生态与真实表达": {
        "description": "围绕社区公约、原创转载、剧情演绎声明、AI 合成标注和理性表达等内容侧规则。",
        "sort_order": 1,
    },
    "广告投放与营销合规": {
        "description": "围绕高危行业营销、投广驳回、行业资质、保证性承诺、产品功效宣传等广告规则。",
        "sort_order": 2,
    },
    "商品交易与经营秩序": {
        "description": "围绕禁售商品、假货山寨、抽奖诈骗、平台经营工具与站内交易秩序。",
        "sort_order": 3,
    },
    "评论与社区互动治理": {
        "description": "围绕评论区广告、不友善互动、站队对立和友好互动治理。",
        "sort_order": 4,
    },
    "未成年人保护": {
        "description": "围绕未成年人不当行为、医美、纹身、化妆、危险场景和模仿风险。",
        "sort_order": 5,
    },
    "运营公告与互动活动": {
        "description": "围绕规则账号的活动运营、粉丝互动、答疑机制和内容预告。",
        "sort_order": 6,
    },
}


NOTE_METADATA: dict[int, dict[str, str]] = {
    1: {
        "category": "内容生态与真实表达",
        "topic": "社区公约 2.0",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": (
            "社区公约 2.0 总纲，拆成真诚分享、友好互动、有序经营三部分，覆盖反对制造对立、"
            "抵制虚假内容、亮明身份诚信经营等核心原则。"
        ),
        "rule_highlights": (
            "AI 辅助创作主动标明；禁止虚假人设、裸露擦边、夸张猎奇标题、"
            "无资质医疗/投资/购房建议、人肉网暴和恶意竞争。"
        ),
    },
    2: {
        "category": "运营公告与互动活动",
        "topic": "规则互动问答",
        "source_value": "规则服务",
        "ingestion_recommendation": "作为旁证",
        "summary": "百篇纪念互动问答，复盘原创素材使用、封建迷信、交易导流和违规案例识别，属于规则普及型内容。",
        "rule_highlights": "原创素材不可抄袭搬运；封建迷信内容违规；站外导流和小号私联违规；通过答题形式做规则教育。",
    },
    3: {
        "category": "商品交易与经营秩序",
        "topic": "禁售商品 / 山寨假货",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "禁售商品系列第 4 篇，明确平台禁止售卖山寨假货、品牌同款擦边货和侵犯知识产权商品。",
        "rule_highlights": (
            "假冒商标、山寨同款、高仿纯原、遮挡 logo 规避审核、蹭品牌热度、"
            "分装品牌化妆品、打板图等都属于高风险。"
        ),
    },
    4: {
        "category": "广告投放与营销合规",
        "topic": "KOS 无资质营销",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": (
            "解读利用低门槛 KOS 身份规避审核、开展高危行业营销的违规场景，"
            "重点覆盖医疗、法律、金融、房地产等行业。"
        ),
        "rule_highlights": (
            "无资质发布疾病诊断/投资建议/教育培训/法律代理内容违规；"
            "百货号包装代查手机号、假货售卖和站外导流都是典型案例。"
        ),
    },
    5: {
        "category": "商品交易与经营秩序",
        "topic": "平台规则动态汇总",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "汇总近期平台规则更新，包含禁售商品新增、蒲公英审核规则和电子资源商家转个人售卖后的合规提示。",
        "rule_highlights": (
            "新增恐怖/极端主义、儿童软色情、AI 魔改历史、劣迹艺人洗白、"
            "明星打榜售卖等禁售内容；强调平台内交易与自动发货合规路径。"
        ),
    },
    6: {
        "category": "广告投放与营销合规",
        "topic": "投广审核 / 商业秩序",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "聚焦商业秩序类广告审核规则，梳理高频驳回点和投放前自查清单。",
        "rule_highlights": (
            "禁止夸张用语、绝对化用语、隐私泄露、制造焦虑、新闻式变相广告、"
            "第三方贬低拉踩和缺乏证明材料的背书。"
        ),
    },
    7: {
        "category": "运营公告与互动活动",
        "topic": "粉丝互动 / 复工话题",
        "source_value": "运营互动",
        "ingestion_recommendation": "仅运营参考",
        "summary": "节后复工互动贴，主要用于粉丝群互动和抽奖活动引流，不承载具体规则条文。",
        "rule_highlights": "无直接规则要点，保留为账号运营语气和互动风格参考。",
    },
    8: {
        "category": "运营公告与互动活动",
        "topic": "春节活动公告",
        "source_value": "运营互动",
        "ingestion_recommendation": "仅运营参考",
        "summary": "春节福利抽奖活动公告，说明活动时间、参与方式和连续更新安排。",
        "rule_highlights": "无直接规则条文，适合作为活动运营和粉丝互动参考，不建议直接入规则库。",
    },
    9: {
        "category": "内容生态与真实表达",
        "topic": "剧情演绎声明",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "专门解释剧情演绎类笔记的声明规范，避免观众将编排内容误认为真实事件。",
        "rule_highlights": (
            "优先使用创作者声明功能；可在字幕、正文话题、资料页、标题标注"
            "“虚构演绎/剧情演绎/内容纯属虚构”等话术。"
        ),
    },
    10: {
        "category": "广告投放与营销合规",
        "topic": "婚恋行业违规套路",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "拆解婚恋行业常见违规套路，包括虚假人设、换软件聊天、投资杀猪盘和无资质婚恋服务。",
        "rule_highlights": "冒充军警、封闭管控借口、外部软件导流、投资收益截图诱骗、无资质婚恋服务推广都属于高风险。",
    },
    11: {
        "category": "广告投放与营销合规",
        "topic": "投广驳回 / 行业认证不符",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "面向二手/二奢行业广告主，解释高危行业在投广时因认证行业不符导致驳回的常见原因。",
        "rule_highlights": "未认证行业推广高危内容、认证主体和推广内容不匹配、跨一级行业投放二手奢品都会触发驳回。",
    },
    12: {
        "category": "内容生态与真实表达",
        "topic": "友好互动 / 反对对立",
        "source_value": "规则服务",
        "ingestion_recommendation": "作为旁证",
        "summary": "转载公益短片，讨论互联网中被曲解、被站队审判的体验，指向社区公约中的反对制造对立与理性对话。",
        "rule_highlights": "强调尊重不同、减少站队和攻击性表达，可作为友好互动治理的传播性补充材料。",
    },
    13: {
        "category": "商品交易与经营秩序",
        "topic": "禁售商品 / 作弊造假",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "禁售商品系列第 3 篇，集中解释作弊造假类商品和服务，包括假证明、学术作弊和违规技术工具。",
        "rule_highlights": (
            "诊断书/学位/实习证明造假、代刷网课、岗位职称买卖、外挂改机、"
            "刷量服务、赔付教程和假装上班服务都被禁售。"
        ),
    },
    14: {
        "category": "广告投放与营销合规",
        "topic": "房地产资质规范",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "围绕房地产行业投广的高危资质要求，解释不同开户身份与所需材料。",
        "rule_highlights": (
            "房地产综合服务平台、开发商、销售代理、物业服务、中介门店等主体"
            "需与认证资质匹配，否则广告被拦截。"
        ),
    },
    15: {
        "category": "评论与社区互动治理",
        "topic": "评论审核 / 商品服务营销",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "评论审核系列第 2 篇，针对评论区广告、素人暗广和水军配合推广做专项治理说明。",
        "rule_highlights": "批量同质化广告、消费者口吻暗广、一唱一和问答式推广都会被视为评论区违规营销。",
    },
    16: {
        "category": "内容生态与真实表达",
        "topic": "转载声明",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "解释如何设置有效的转载声明，并澄清“礼貌拿文案”“只截切片不算搬运”等常见误区。",
        "rule_highlights": "只要内容任一部分源自他人作品就需要声明转载；转载声明需清晰、有效且可识别。",
    },
    17: {
        "category": "广告投放与营销合规",
        "topic": "房地产保证性承诺",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "聚焦房地产行业的保证性承诺雷区，梳理最常见的承诺式违规表达。",
        "rule_highlights": (
            "不得承诺地理位置、落户升学、投资回报、未确定规划和预售样板间效果；"
            "避免“黄金地段”“稳赚不赔”等说法。"
        ),
    },
    18: {
        "category": "广告投放与营销合规",
        "topic": "广告保证性承诺",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "通用广告规则，禁止对产品或服务效果做出明示或暗示的保证性承诺，并给出教育和医疗行业案例。",
        "rule_highlights": "如“100%有效”“绝对无误”“全部满意”“无条件退费”等都是典型违规表达。",
    },
    19: {
        "category": "商品交易与经营秩序",
        "topic": "禁售商品 / 违禁物品",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "禁售商品系列第 2 篇，聚焦违禁物品、破坏生态、博彩、烟草、侵犯隐私和违规金融商品。",
        "rule_highlights": "野生动植物制品、赌博器具、烟草电子烟、偷窥窃听设备、违规荐股/换汇等都属于禁售范围。",
    },
    20: {
        "category": "评论与社区互动治理",
        "topic": "评论审核 / 不友善互动",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "评论审核系列第 1 篇，解释不友善互动的审核范围、处罚方式和典型违规表达。",
        "rule_highlights": "辱骂、人身攻击、诅咒威胁、低俗歧视和图片评论中的攻击性文字都在治理范围内。",
    },
    21: {
        "category": "广告投放与营销合规",
        "topic": "投广 / 超出产品功效范围",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "从家居百货和母婴用品案例出发，解释广告中不得超出产品真实功效范围。",
        "rule_highlights": "非医疗产品不得宣称治病、医美或保健作用；也不能借用医院、药房等专业渠道做背书。",
    },
    22: {
        "category": "广告投放与营销合规",
        "topic": "投广 / 超出商品功效范围",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "延续功效范围主题，换成食品饮料和美妆个护案例，方便不同类目广告主对照自查。",
        "rule_highlights": "禁止治疗功效、保健功效和医美类联想宣传，所有功效表述必须严格落在产品真实功能上。",
    },
    23: {
        "category": "商品交易与经营秩序",
        "topic": "禁售商品 / 危险管控物品",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "禁售商品系列第 1 篇，聚焦危险/管控类物品和对人身安全、公共秩序有威胁的商品。",
        "rule_highlights": "军火、枪支弹药、仿真枪、管制刀具、电击器、防狼喷雾、麻醉针、毒品及吸毒工具均严禁售卖。",
    },
    24: {
        "category": "内容生态与真实表达",
        "topic": "AI 内容风险",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "系统梳理 AI 内容风险，包括冒充名人/他人、军警形象篡改、虚假信息、低俗养号和 AI 广告营销。",
        "rule_highlights": (
            "AI 冒充名人、AI 虚假人设、AI 虚假灾害新闻、AI 低俗内容、"
            "AI 骗互动和 AI 广告营销都需重点管控；建议主动声明 AI 合成内容。"
        ),
    },
    25: {
        "category": "商品交易与经营秩序",
        "topic": "官方经营工具",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "对比个人售卖、获客工具和小红书店铺三类官方经营工具的适用人群、门槛和费用。",
        "rule_highlights": (
            "强调通过官方工具完成交易或留资才是合规路径；个人售卖、企业专业号"
            "和聚光消耗门槛分别对应不同工具。"
        ),
    },
    26: {
        "category": "运营公告与互动活动",
        "topic": "官方违规答疑",
        "source_value": "规则服务",
        "ingestion_recommendation": "作为旁证",
        "summary": "官方答疑入口型笔记，教用户如何附带违规笔记链接、处置详情与问题描述进行咨询。",
        "rule_highlights": "适合作为申诉和咨询流程参考，不是规则条文本体，但能补足用户服务流程信息。",
    },
    27: {
        "category": "运营公告与互动活动",
        "topic": "账号阶段总结",
        "source_value": "运营互动",
        "ingestion_recommendation": "仅运营参考",
        "summary": "规则百科薯 2025 到 2026 的运营总结和下一步计划，体现账号定位与内容方向。",
        "rule_highlights": "无直接规则条文，可用于理解账号后续会围绕新规速递、评论答疑和误伤处理持续更新。",
    },
    28: {
        "category": "商品交易与经营秩序",
        "topic": "抽奖送礼防骗",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "围绕抽奖/免费送场景的诈骗风险与合规边界，帮助识别假冒商家和导流陷阱。",
        "rule_highlights": "假冒商家抽奖、无背景实体抽奖、无理由送虚拟物品、虚假“人人有份”和借抽奖导流都属于高风险。",
    },
    29: {
        "category": "未成年人保护",
        "topic": "未成年不当行为（二）",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "未成年规则第二期，覆盖未成年医美、化妆、佩戴穿刺物和吃播等容易引发模仿或安全风险的内容。",
        "rule_highlights": "未成年人咨询/展示医美、以学生优惠诱导医美、展示化妆过程或不良画风妆容等均在限制范围内。",
    },
    30: {
        "category": "未成年人保护",
        "topic": "未成年不当行为（一）",
        "source_value": "核心规则",
        "ingestion_recommendation": "优先入库",
        "summary": "未成年规则第一期，聚焦纹身、喊麦、社会摇、不当表达和出入不当场所。",
        "rule_highlights": "未成年人展示纹身、喊麦、社会摇、说脏话、出入 KTV 等内容都属于重点治理对象。",
    },
    31: {
        "category": "运营公告与互动活动",
        "topic": "内容预告",
        "source_value": "运营互动",
        "ingestion_recommendation": "仅运营参考",
        "summary": "保暖问候加内容预告，提前告知后续会围绕广告投放、商家避坑和行业话题持续更新。",
        "rule_highlights": "无直接规则条文，可用于理解账号早期的内容规划和运营表达。",
    },
}


THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(value.replace("\n", " ").split())


def truncate(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def to_date_text(timestamp_ms: int | None) -> str:
    if not timestamp_ms:
        return ""
    return datetime.fromtimestamp(timestamp_ms / 1000).strftime("%Y-%m-%d")


def build_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    notes = read_json(NOTES_INDEX_PATH)
    rows: list[dict[str, Any]] = []
    image_rows: list[dict[str, Any]] = []

    for note in notes:
        index = int(note["index"])
        meta = NOTE_METADATA[index]
        ocr_path = OCR_DIR / f"{index:03d}_{note['note_id']}.json"
        ocr_data = read_json(ocr_path)
        full_text = normalize_text((ocr_data.get("ocr") or {}).get("full_text"))
        desc = normalize_text(note.get("desc"))
        tags = " / ".join(note.get("tags") or [])

        row = {
            "序号": index,
            "发布时间": to_date_text(note.get("published_time")),
            "标题": note.get("title", ""),
            "一级分类": meta["category"],
            "二级主题": meta["topic"],
            "来源价值": meta["source_value"],
            "入库建议": meta["ingestion_recommendation"],
            "内容摘要": meta["summary"],
            "核心规则点": meta["rule_highlights"],
            "原始正文摘录": truncate(desc, 220),
            "图片数": note.get("image_count", 0),
            "图片文字要点": truncate(full_text, 260),
            "标签": tags,
            "笔记链接": note.get("note_url", ""),
            "原始 HTML": note.get("html_path", ""),
            "原始 JSON": str(MERGED_DIR / "notes" / f"{index:03d}_{note['note_id']}.json"),
            "OCR JSON": str(ocr_path),
        }
        rows.append(row)

        image_paths = note.get("images") or []
        ocr_images = (ocr_data.get("ocr") or {}).get("images") or []
        for image_index, image_path in enumerate(image_paths, start=1):
            ocr_text = ""
            confidence_avg = ""
            if image_index - 1 < len(ocr_images):
                image_info = ocr_images[image_index - 1]
                ocr_text = normalize_text(image_info.get("text"))
                lines = image_info.get("lines") or []
                if lines:
                    confidence_avg = round(
                        sum(float(line.get("confidence", 0)) for line in lines) / len(lines),
                        3,
                    )
            image_rows.append(
                {
                    "序号": index,
                    "标题": note.get("title", ""),
                    "一级分类": meta["category"],
                    "二级主题": meta["topic"],
                    "图片序号": image_index,
                    "图片路径": image_path,
                    "OCR 置信度均值": confidence_avg,
                    "OCR 文本": ocr_text,
                }
            )

    rows.sort(key=lambda item: item["序号"])
    image_rows.sort(key=lambda item: (item["序号"], item["图片序号"]))
    return rows, image_rows


def style_header(worksheet, row_index: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[row_index]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_table_style(worksheet, wrap_columns: set[int] | None = None) -> None:
    wrap_columns = wrap_columns or set()
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=cell.column in wrap_columns,
            )


def auto_width(worksheet, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def fill_sheet(worksheet, rows: list[dict[str, Any]]) -> None:
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    style_header(worksheet)
    apply_table_style(worksheet, wrap_columns={4, 8, 9, 10, 12, 13})
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    auto_width(
        worksheet,
        {
            1: 8,
            2: 12,
            3: 28,
            4: 18,
            5: 18,
            6: 12,
            7: 12,
            8: 32,
            9: 44,
            10: 34,
            11: 8,
            12: 44,
            13: 22,
            14: 48,
            15: 44,
            16: 44,
            17: 44,
        },
    )


def build_workbook(rows: list[dict[str, Any]], image_rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()

    overview = workbook.active
    overview.title = "分类总览"
    overview_headers = [
        "一级分类",
        "分类说明",
        "笔记数",
        "核心规则数",
        "规则服务数",
        "运营互动数",
        "序号列表",
        "笔记标题列表",
    ]
    overview.append(overview_headers)
    categories = sorted(
        CATEGORY_INFO.items(),
        key=lambda item: item[1]["sort_order"],
    )
    for category_name, category_info in categories:
        category_rows = [row for row in rows if row["一级分类"] == category_name]
        overview.append(
            [
                category_name,
                category_info["description"],
                len(category_rows),
                sum(1 for row in category_rows if row["来源价值"] == "核心规则"),
                sum(1 for row in category_rows if row["来源价值"] == "规则服务"),
                sum(1 for row in category_rows if row["来源价值"] == "运营互动"),
                ", ".join(f"{row['序号']:03d}" for row in category_rows),
                " / ".join(row["标题"] for row in category_rows),
            ]
        )
    style_header(overview)
    apply_table_style(overview, wrap_columns={2, 8})
    overview.freeze_panes = "A2"
    auto_width(
        overview,
        {
            1: 18,
            2: 42,
            3: 10,
            4: 10,
            5: 10,
            6: 10,
            7: 18,
            8: 72,
        },
    )

    notes_sheet = workbook.create_sheet("笔记总表")
    fill_sheet(notes_sheet, rows)

    image_sheet = workbook.create_sheet("图片OCR明细")
    image_headers = list(image_rows[0].keys())
    image_sheet.append(image_headers)
    for row in image_rows:
        image_sheet.append([row.get(header, "") for header in image_headers])
    style_header(image_sheet)
    apply_table_style(image_sheet, wrap_columns={3, 4, 8})
    image_sheet.freeze_panes = "A2"
    image_sheet.auto_filter.ref = image_sheet.dimensions
    auto_width(
        image_sheet,
        {
            1: 8,
            2: 28,
            3: 18,
            4: 18,
            5: 10,
            6: 52,
            7: 14,
            8: 56,
        },
    )

    for category_name, _category_info in categories:
        category_sheet = workbook.create_sheet(category_name)
        category_rows = [row for row in rows if row["一级分类"] == category_name]
        fill_sheet(category_sheet, category_rows)

    source_sheet = workbook.create_sheet("来源说明")
    source_sheet.append(["项目", "内容"])
    source_sheet.append(["来源主页", "https://www.xiaohongshu.com/user/profile/677e62aa000000000801e7c3"])
    source_sheet.append(["抓取对象", "规则百科薯账号前 31 篇顺序笔记"])
    source_sheet.append(["正文索引", str(NOTES_INDEX_PATH)])
    source_sheet.append(["OCR 目录", str(OCR_DIR)])
    source_sheet.append(["输出文件", str(OUTPUT_PATH)])
    source_sheet.append(
        [
            "整理方法",
            "按单篇笔记抓取正文与图片，再对图片做 OCR，最后按主题分类并补充每篇摘要、入库建议与规则点。",
        ]
    )
    source_sheet.append(
        [
            "入库建议说明",
            "优先入库=直接规则条文或规则解读；作为旁证=答疑/传播型内容；仅运营参考=活动公告或运营互动。",
        ]
    )
    style_header(source_sheet)
    apply_table_style(source_sheet, wrap_columns={2})
    auto_width(source_sheet, {1: 18, 2: 88})

    return workbook


def main() -> None:
    rows, image_rows = build_rows()
    workbook = build_workbook(rows, image_rows)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
