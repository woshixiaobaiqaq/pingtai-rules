from __future__ import annotations

import argparse
import datetime as dt
import html
import json
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def parse_timestamp(timestamp: int | None) -> str:
    if not timestamp:
        return ""
    return dt.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")


def collect_table_image_ids(deltas: dict[str, Any], zone_ref: str, seen: set[str] | None = None) -> list[str]:
    seen = seen or set()
    image_ids: list[str] = []
    for zone_id in zone_ref.split():
        if zone_id in seen or zone_id not in deltas:
            continue
        seen.add(zone_id)
        for op in deltas[zone_id].get("ops", []):
            attrs = op.get("attributes", {})
            if attrs.get("IMAGE") == "true" and attrs.get("uri"):
                image_ids.append(attrs["uri"])
            if "aceTable" in attrs:
                image_ids.extend(collect_table_image_ids(deltas, attrs["aceTable"], seen))
            inserted = op.get("insert")
            if isinstance(inserted, dict) and "id" in inserted:
                image_ids.extend(collect_table_image_ids(deltas, inserted["id"], seen))
    return image_ids


def collect_table_image_urls(deltas: dict[str, Any], zone_ref: str, seen: set[str] | None = None) -> list[str]:
    seen = seen or set()
    image_urls: list[str] = []
    for zone_id in zone_ref.split():
        if zone_id in seen or zone_id not in deltas:
            continue
        seen.add(zone_id)
        for op in deltas[zone_id].get("ops", []):
            attrs = op.get("attributes", {})
            if attrs.get("IMAGE") == "true" and attrs.get("fileSrc"):
                image_urls.append(attrs["fileSrc"])
            if "aceTable" in attrs:
                image_urls.extend(collect_table_image_urls(deltas, attrs["aceTable"], seen))
            inserted = op.get("insert")
            if isinstance(inserted, dict) and "id" in inserted:
                image_urls.extend(collect_table_image_urls(deltas, inserted["id"], seen))
    return image_urls


def extract_all_image_urls(content_data: dict[str, Any]) -> list[str]:
    deltas = content_data["deltas"]
    urls = [
        attachment["url"] for attachment in content_data.get("attachments", []) if attachment.get("type") == "image"
    ]
    for op in deltas["0"].get("ops", []):
        attrs = op.get("attributes", {})
        if attrs.get("IMAGE") == "true" and attrs.get("fileSrc"):
            urls.append(attrs["fileSrc"])
        if "aceTable" in attrs:
            urls.extend(collect_table_image_urls(deltas, attrs["aceTable"]))

    deduped: list[str] = []
    seen: set[str] = set()
    for url in urls:
        if url not in seen:
            deduped.append(url)
            seen.add(url)
    return deduped


def image_uri_from_remote_url(url: str) -> str:
    parsed = urlparse(url)
    path = parsed.path.strip("/")
    if "~" in path:
        path = path.split("~", 1)[0]
    return path


def format_inline_text(text: str, attrs: dict[str, Any]) -> str:
    value = html.unescape(text).strip()
    if not value:
        return ""
    if attrs.get("bold") == "true":
        return value
    return value


def parse_blocks(content_data: dict[str, Any], image_map: dict[str, str]) -> list[dict[str, Any]]:
    deltas = content_data["deltas"]
    root_ops = deltas["0"]["ops"]
    blocks: list[dict[str, Any]] = []
    current_fragments: list[str] = []
    pending_attrs: dict[str, Any] = {}
    current_section = ""
    current_group = ""
    block_order = 0

    def flush_text_block() -> None:
        nonlocal current_fragments, pending_attrs, current_section, current_group, block_order
        if not current_fragments:
            pending_attrs = {}
            return

        text = "".join(current_fragments).strip()
        current_fragments = []
        if not text:
            pending_attrs = {}
            return

        block_type = "paragraph"
        if pending_attrs.get("heading") == "h3":
            block_type = "section_heading"
            current_section = text
            current_group = ""
        elif pending_attrs.get("blockquote") == "true":
            block_type = "faq"
        elif pending_attrs.get("list"):
            block_type = "bullet"
        elif text in {"常见违规点", "常见问题"}:
            block_type = "group_heading"
            current_group = text

        block_order += 1
        blocks.append(
            {
                "block_order": block_order,
                "block_type": block_type,
                "section_heading": current_section,
                "group_heading": current_group,
                "text": text,
                "image_refs": [],
            }
        )
        pending_attrs = {}

    def append_image_block(image_ids: list[str]) -> None:
        nonlocal block_order
        if not image_ids:
            return
        block_order += 1
        blocks.append(
            {
                "block_order": block_order,
                "block_type": "image",
                "section_heading": current_section,
                "group_heading": current_group,
                "text": "",
                "image_refs": [image_map.get(image_id, image_id) for image_id in image_ids],
            }
        )

    for op in root_ops:
        attrs = op.get("attributes", {})
        inserted = op.get("insert")

        if inserted == "*" and attrs.get("lmkr") == "1":
            if any(key in attrs for key in ("heading", "list", "blockquote")):
                pending_attrs = attrs
            continue

        if attrs.get("IMAGE") == "true" and attrs.get("uri"):
            flush_text_block()
            append_image_block([attrs["uri"]])
            continue

        if "aceTable" in attrs:
            flush_text_block()
            append_image_block(collect_table_image_ids(deltas, attrs["aceTable"]))
            continue

        if isinstance(inserted, dict):
            continue
        if not isinstance(inserted, str):
            continue

        parts = inserted.split("\n")
        for index, part in enumerate(parts):
            piece = format_inline_text(part, attrs)
            if piece:
                current_fragments.append(piece)
            if index < len(parts) - 1:
                flush_text_block()

    flush_text_block()
    return blocks


def apply_sheet_style(sheet, widths: dict[str, float], freeze_cell: str) -> None:
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(bold=True, color="1F1F1F")
    thin = Side(style="thin", color="D9D9D9")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = freeze_cell
    sheet.auto_filter.ref = sheet.dimensions


def export_archive_to_excel(archive_dir: Path, output_path: Path) -> Path:
    manifest = load_json(archive_dir / "archive_manifest.json")
    workbook = Workbook()

    overview_sheet = workbook.active
    overview_sheet.title = "分类总览"
    overview_headers = [
        "标题",
        "commerce_id",
        "描述",
        "创建时间",
        "更新时间",
        "浏览量",
        "图片数量",
        "原始 JSON",
        "离线页面",
        "线上页面",
    ]
    overview_sheet.append(overview_headers)

    block_sheet = workbook.create_sheet("正文块明细")
    block_headers = [
        "标题",
        "commerce_id",
        "块序号",
        "块类型",
        "章节标题",
        "分组标题",
        "正文内容",
        "图片引用",
    ]
    block_sheet.append(block_headers)

    image_sheet = workbook.create_sheet("图片清单")
    image_headers = [
        "标题",
        "commerce_id",
        "图片序号",
        "本地路径",
        "远程 URL",
        "文件大小(字节)",
    ]
    image_sheet.append(image_headers)

    for article in manifest["articles"]:
        article_json_path = archive_dir / article["article_json_path"]
        article_payload = load_json(article_json_path)
        article_info = article_payload["data"]["article_info"]
        content_data = json.loads(article_info["content"])

        image_map: dict[str, str] = {}
        remote_image_urls = extract_all_image_urls(content_data)
        for image_index, (image_path, remote_url) in enumerate(
            zip(article["image_paths"], remote_image_urls, strict=False),
            start=1,
        ):
            image_file = archive_dir / image_path
            image_map[image_uri_from_remote_url(remote_url)] = image_path
            image_sheet.append(
                [
                    article["title"],
                    article["commerce_id"],
                    image_index,
                    image_path,
                    remote_url,
                    image_file.stat().st_size if image_file.exists() else 0,
                ]
            )

        overview_sheet.append(
            [
                article["title"],
                article["commerce_id"],
                article_info.get("description", ""),
                parse_timestamp(article_info.get("create_timestamp")),
                parse_timestamp(article_info.get("update_timestamp")),
                article_info.get("view_count", 0),
                len(article["image_paths"]),
                article["article_json_path"],
                article["rendered_html_path"],
                article["source_url"],
            ]
        )

        blocks = parse_blocks(content_data, image_map)
        for block in blocks:
            block_sheet.append(
                [
                    article["title"],
                    article["commerce_id"],
                    block["block_order"],
                    block["block_type"],
                    block["section_heading"],
                    block["group_heading"],
                    block["text"],
                    "\n".join(block["image_refs"]),
                ]
            )

    for sheet in (overview_sheet, block_sheet, image_sheet):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    apply_sheet_style(
        overview_sheet,
        {
            "A": 22,
            "B": 18,
            "C": 56,
            "D": 20,
            "E": 20,
            "F": 12,
            "G": 10,
            "H": 34,
            "I": 28,
            "J": 42,
        },
        "A2",
    )
    apply_sheet_style(
        block_sheet,
        {
            "A": 18,
            "B": 18,
            "C": 10,
            "D": 14,
            "E": 20,
            "F": 14,
            "G": 72,
            "H": 40,
        },
        "A2",
    )
    apply_sheet_style(
        image_sheet,
        {
            "A": 18,
            "B": 18,
            "C": 10,
            "D": 44,
            "E": 64,
            "F": 14,
        },
        "A2",
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Export the archived Douyin rule explain data to Excel.")
    parser.add_argument(
        "--archive-dir",
        type=Path,
        default=Path("data/source_archives/douyin_rule_explain/full_archive_2026-03-11"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("output/spreadsheet/douyin_rule_explain_full_archive_2026-03-11.xlsx"),
    )
    args = parser.parse_args()
    result = export_archive_to_excel(args.archive_dir.resolve(), args.output.resolve())
    print(result)


if __name__ == "__main__":
    main()
