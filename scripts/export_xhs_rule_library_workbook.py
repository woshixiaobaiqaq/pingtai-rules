from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/m/亿一的AI小助理/open-codex/platform-content-audit")
RULES_PATH = ROOT / "data/rule_library/xiaohongshu/rules.json"
CATALOG_PATH = ROOT / "data/rule_library/xiaohongshu/catalog.json"
MANIFEST_PATH = ROOT / "data/rule_library/manifest.json"
OUTPUT_DIR = ROOT / "output/spreadsheet"
OUTPUT_PATH = OUTPUT_DIR / "xiaohongshu_rule_library_complete_2026-03-16.xlsx"

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(value.replace("\n", " ").split())


def join_values(values: list[str] | tuple[str, ...] | None) -> str:
    if not values:
        return ""
    return " / ".join(str(value) for value in values if value)


def truncate(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def derive_primary_category(rule: dict[str, Any]) -> str:
    metadata = rule.get("metadata", {})
    if metadata.get("manual_category"):
        return str(metadata["manual_category"])
    if metadata.get("pillar"):
        return str(metadata["pillar"])
    title = str(rule.get("title", ""))
    if "：" in title:
        return title.split("：", 1)[0]
    return "未分类"


def derive_topic(rule: dict[str, Any]) -> str:
    metadata = rule.get("metadata", {})
    if metadata.get("manual_title"):
        return str(metadata["manual_title"])
    title = str(rule.get("title", ""))
    if "：" in title:
        return title.split("：", 1)[1]
    return title


def derive_source_group(rule: dict[str, Any]) -> str:
    metadata = rule.get("metadata", {})
    if metadata.get("source_type") == "creator_center_manual_capture":
        return "创作中心截图规则"
    return "公开资料主题规则"


def derive_source_urls(rule: dict[str, Any]) -> list[str]:
    metadata = rule.get("metadata", {})
    urls = metadata.get("source_urls") or []
    if isinstance(urls, list) and urls:
        return [str(url) for url in urls if url]
    source_url = rule.get("source_url")
    return [str(source_url)] if source_url else []


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_table_style(sheet, wrap_columns: set[int]) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in wrap_columns)


def set_widths(sheet, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def build_overview_sheet(workbook: Workbook, rules: list[dict[str, Any]]) -> None:
    sheet = workbook.active
    sheet.title = "库总览"
    sheet.append(["指标", "数值"])

    severity_counter = Counter(str(rule.get("severity", "")) for rule in rules)
    source_counter = Counter(derive_source_group(rule) for rule in rules)
    manual_categories = {
        derive_primary_category(rule)
        for rule in rules
        if derive_source_group(rule) == "创作中心截图规则"
    }
    public_categories = {
        derive_primary_category(rule)
        for rule in rules
        if derive_source_group(rule) == "公开资料主题规则"
    }

    rows = [
        ("生成时间", datetime.now().isoformat(timespec="seconds")),
        ("规则总数", len(rules)),
        ("公开资料主题规则", source_counter["公开资料主题规则"]),
        ("创作中心截图规则", source_counter["创作中心截图规则"]),
        ("高风险规则", severity_counter["high"]),
        ("中风险规则", severity_counter["medium"]),
        ("低风险规则", severity_counter["low"]),
        ("截图规则一级分类数", len(manual_categories)),
        ("公开资料主题分类数", len(public_categories)),
        ("规则文件", str(RULES_PATH)),
        ("目录文件", str(CATALOG_PATH)),
    ]
    for row in rows:
        sheet.append(list(row))

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={2})
    set_widths(sheet, {1: 24, 2: 120})


def build_rule_table_sheet(workbook: Workbook, rules: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("完整规则总表")
    sheet.append(
        [
            "序号",
            "规则ID",
            "平台",
            "来源分组",
            "一级分类",
            "二级主题",
            "标题",
            "风险等级",
            "规则正文",
            "关键词",
            "正则",
            "标签",
            "推断候选标签",
            "主来源",
            "全部来源",
            "OCR标题来源",
            "OCR置信度",
            "异常标记",
        ]
    )

    for index, rule in enumerate(rules, start=1):
        metadata = rule.get("metadata", {})
        source_urls = derive_source_urls(rule)
        sheet.append(
            [
                index,
                rule.get("rule_id", ""),
                rule.get("platform", ""),
                derive_source_group(rule),
                derive_primary_category(rule),
                derive_topic(rule),
                rule.get("title", ""),
                rule.get("severity", ""),
                normalize_text(rule.get("content")),
                join_values(rule.get("keywords")),
                join_values(rule.get("regex_patterns")),
                join_values(rule.get("tags")),
                join_values(metadata.get("inferred_candidate_tags")),
                rule.get("source_url") or "",
                join_values(source_urls),
                metadata.get("title_source", ""),
                metadata.get("ocr_confidence_avg", ""),
                join_values(metadata.get("flags")),
            ]
        )

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={6, 7, 9, 10, 11, 12, 13, 14, 15, 18})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(
        sheet,
        {
            1: 8,
            2: 16,
            3: 12,
            4: 16,
            5: 18,
            6: 24,
            7: 28,
            8: 10,
            9: 88,
            10: 28,
            11: 28,
            12: 24,
            13: 24,
            14: 28,
            15: 52,
            16: 14,
            17: 12,
            18: 18,
        },
    )


def build_category_summary_sheet(workbook: Workbook, rules: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("分类汇总")
    sheet.append(
        [
            "一级分类",
            "来源分组",
            "规则数",
            "高风险",
            "中风险",
            "低风险",
            "示例条目",
        ]
    )

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for rule in rules:
        key = (derive_primary_category(rule), derive_source_group(rule))
        grouped[key].append(rule)

    for (category, source_group), items in sorted(grouped.items()):
        severity_counter = Counter(str(item.get("severity", "")) for item in items)
        sample_titles = " / ".join(derive_topic(item) for item in items[:6])
        sheet.append(
            [
                category,
                source_group,
                len(items),
                severity_counter["high"],
                severity_counter["medium"],
                severity_counter["low"],
                sample_titles,
            ]
        )

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={7})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(sheet, {1: 18, 2: 16, 3: 10, 4: 10, 5: 10, 6: 10, 7: 88})


def build_manual_detail_sheet(workbook: Workbook, rules: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("截图规则明细")
    sheet.append(
        [
            "规则ID",
            "一级分类",
            "条款标题",
            "风险等级",
            "关键词",
            "图片文件",
            "图片路径",
            "OCR标题来源",
            "OCR置信度",
            "异常标记",
            "OCR正文摘录",
        ]
    )

    manual_rules = [rule for rule in rules if derive_source_group(rule) == "创作中心截图规则"]
    for rule in manual_rules:
        metadata = rule.get("metadata", {})
        source_urls = derive_source_urls(rule)
        sheet.append(
            [
                rule.get("rule_id", ""),
                derive_primary_category(rule),
                derive_topic(rule),
                rule.get("severity", ""),
                join_values(rule.get("keywords")),
                metadata.get("image_name", ""),
                source_urls[0] if source_urls else "",
                metadata.get("title_source", ""),
                metadata.get("ocr_confidence_avg", ""),
                join_values(metadata.get("flags")),
                truncate(normalize_text(rule.get("content")), 260),
            ]
        )

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={5, 7, 10, 11})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(sheet, {1: 14, 2: 16, 3: 24, 4: 10, 5: 24, 6: 20, 7: 62, 8: 14, 9: 12, 10: 18, 11: 82})


def build_source_sheet(workbook: Workbook, manifest: dict[str, Any], catalog: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("来源说明")
    sheet.append(["项目", "内容"])
    sheet.append(["规则清单", str(RULES_PATH)])
    sheet.append(["目录清单", str(CATALOG_PATH)])
    sheet.append(["规则库版本", str(manifest.get("version", ""))])
    sheet.append(["小红书目录版本", str(catalog.get("version", ""))])
    sheet.append(["范围", str(catalog.get("scope", ""))])

    for index, item in enumerate(catalog.get("source_notes", []), start=1):
        sheet.append([f"来源 {index}", json.dumps(item, ensure_ascii=False)])

    style_header(sheet)
    apply_table_style(sheet, wrap_columns={2})
    set_widths(sheet, {1: 16, 2: 120})


def main() -> None:
    rules_bundle = read_json(RULES_PATH)
    catalog = read_json(CATALOG_PATH)
    manifest = read_json(MANIFEST_PATH)
    rules = list(rules_bundle.get("rules", []))

    workbook = Workbook()
    build_overview_sheet(workbook, rules)
    build_rule_table_sheet(workbook, rules)
    build_category_summary_sheet(workbook, rules)
    build_manual_detail_sheet(workbook, rules)
    build_source_sheet(workbook, manifest, catalog)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)

    print(
        json.dumps(
            {
                "output_path": str(OUTPUT_PATH),
                "rule_count": len(rules),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
