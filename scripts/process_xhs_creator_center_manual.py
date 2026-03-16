from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidocr_onnxruntime import RapidOCR

ROOT = Path("/Users/m/亿一的AI小助理/open-codex/platform-content-audit")
SOURCE_ROOT = ROOT / "data/source_archives/xiaohongshu_creator_center_manual"
INBOX_DIR = SOURCE_ROOT / "inbox"
OUTPUT_ROOT = SOURCE_ROOT / "processed"
SPREADSHEET_DIR = ROOT / "output/spreadsheet"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".heic", ".webp"}
GENERIC_STEMS = {"screenshot", "screen_shot", "image", "img"}

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


@dataclass(slots=True)
class ImageOcrRecord:
    category_index: int
    category_name: str
    file_stem: str
    display_title: str
    title_source: str
    image_path: str
    image_name: str
    full_text: str
    confidence_avg: float | None
    line_count: int
    flags: list[str]


def list_category_dirs() -> list[Path]:
    return sorted(
        [
            path
            for path in INBOX_DIR.iterdir()
            if path.is_dir() and not path.name.startswith(".")
        ]
    )


def normalize_text(value: str) -> str:
    return " ".join(value.replace("\n", " ").split())


def sanitize_identifier(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = "".join(char for char in normalized if not unicodedata.combining(char))
    ascii_only = ascii_only.lower()
    ascii_only = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "-", ascii_only)
    ascii_only = re.sub(r"-{2,}", "-", ascii_only).strip("-")
    return ascii_only or "unknown"


def split_category_name(dirname: str) -> tuple[int, str]:
    match = re.match(r"^(\d+)[_-](.+)$", dirname)
    if match:
        return int(match.group(1)), match.group(2)
    return 999, dirname


def is_generic_stem(stem: str) -> bool:
    simplified = stem.lower().strip()
    if simplified.startswith("screenshot_"):
        return True
    return simplified in GENERIC_STEMS


def extract_title_from_ocr(category_name: str, full_text: str) -> str | None:
    if not full_text:
        return None

    prefix = full_text.split("规则说明", 1)[0]
    prefix = prefix.replace("\n", " ")
    prefix = prefix.replace("：", " ").replace(":", " ")
    prefix = prefix.replace("—", " ").replace("-", " ")
    prefix = normalize_text(prefix)
    if not prefix:
        return None

    pieces = [piece for piece in re.split(r"\s+", prefix) if piece]
    filtered: list[str] = []
    category_variants = {category_name, category_name.replace("与", ""), category_name.replace("及", "")}
    for piece in pieces:
        if piece in {"规则说明", "规则", "说明"}:
            continue
        if any(piece == variant or piece.startswith(variant) for variant in category_variants):
            continue
        filtered.append(piece)

    if not filtered:
        return None

    # Prefer the last short phrase in the header area because the screenshots often
    # look like "一级类-二级类 关联词 规则标题".
    candidates = filtered[-3:]
    best = max(candidates, key=len)
    best = best.strip("·•|/ ")
    return best or None


def ocr_image(engine: RapidOCR, image_path: Path) -> tuple[str, list[dict[str, float | str]], float | None]:
    result, _ = engine(str(image_path))
    if not result:
        return "", [], None

    lines = []
    confidences: list[float] = []
    for item in result:
        line_text = normalize_text(item[1])
        confidence = float(item[2])
        lines.append(
            {
                "text": line_text,
                "confidence": confidence,
            }
        )
        confidences.append(confidence)

    return (
        " ".join(line["text"] for line in lines),
        lines,
        round(sum(confidences) / len(confidences), 4) if confidences else None,
    )


def style_header(worksheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_table_style(worksheet, wrap_columns: set[int]) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in wrap_columns)


def set_widths(worksheet, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def build_workbook(records: list[ImageOcrRecord], workbook_path: Path) -> None:
    workbook = Workbook()

    overview = workbook.active
    overview.title = "分类总览"
    overview.append(["一级分类", "条目数", "命名异常数", "OCR 低置信度数", "条款标题列表"])

    category_map: dict[str, list[ImageOcrRecord]] = {}
    for record in records:
        category_map.setdefault(record.category_name, []).append(record)

    for category_name, items in sorted(category_map.items(), key=lambda item: item[1][0].category_index):
        naming_issues = sum(1 for item in items if "generic_filename" in item.flags)
        low_confidence = sum(1 for item in items if "low_confidence" in item.flags)
        overview.append(
            [
                category_name,
                len(items),
                naming_issues,
                low_confidence,
                " / ".join(item.display_title for item in items),
            ]
        )

    style_header(overview)
    apply_table_style(overview, wrap_columns={5})
    overview.freeze_panes = "A2"
    set_widths(overview, {1: 18, 2: 10, 3: 12, 4: 14, 5: 88})

    details = workbook.create_sheet("规则条款明细")
    details.append(
        [
            "一级分类序号",
            "一级分类",
            "原文件名",
            "条款标题",
            "标题来源",
            "OCR 置信度",
            "OCR 行数",
            "异常标记",
            "OCR 正文",
            "图片路径",
        ]
    )
    for record in records:
        details.append(
            [
                record.category_index,
                record.category_name,
                record.image_name,
                record.display_title,
                record.title_source,
                record.confidence_avg,
                record.line_count,
                " / ".join(record.flags),
                record.full_text,
                record.image_path,
            ]
        )

    style_header(details)
    apply_table_style(details, wrap_columns={4, 8, 9, 10})
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    set_widths(
        details,
        {
            1: 10,
            2: 16,
            3: 24,
            4: 24,
            5: 12,
            6: 12,
            7: 10,
            8: 18,
            9: 86,
            10: 68,
        },
    )

    source_sheet = workbook.create_sheet("来源说明")
    source_sheet.append(["项目", "内容"])
    source_sheet.append(["来源目录", str(INBOX_DIR)])
    source_sheet.append(["处理时间", datetime.now().isoformat(timespec="seconds")])
    source_sheet.append(
        [
            "处理方式",
            "按一级分类目录扫描图片，逐图 OCR，结合文件名和 OCR 头部推断条款标题。",
        ]
    )
    source_sheet.append(
        [
            "命名规则",
            "目录名=一级大类；图片名=规则条款。若文件名是 screenshot_* 等通用名，"
            "则使用 OCR 标题兜底。",
        ]
    )

    style_header(source_sheet)
    apply_table_style(source_sheet, wrap_columns={2})
    set_widths(source_sheet, {1: 18, 2: 92})

    workbook.save(workbook_path)


def main() -> None:
    categories = list_category_dirs()
    engine = RapidOCR()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_dir = OUTPUT_ROOT / f"processed_{timestamp}"
    output_dir.mkdir(parents=True, exist_ok=True)

    records: list[ImageOcrRecord] = []
    per_category_payload = []

    for category_dir in categories:
        category_index, category_name = split_category_name(category_dir.name)
        items_payload = []
        image_paths = sorted(
            [
                path
                for path in category_dir.iterdir()
                if path.is_file()
                and path.suffix.lower() in IMAGE_EXTENSIONS
                and not path.name.startswith(".")
            ]
        )

        for image_path in image_paths:
            full_text, lines, confidence_avg = ocr_image(engine, image_path)
            file_stem = image_path.stem
            title_source = "filename"
            display_title = file_stem
            flags: list[str] = []

            if is_generic_stem(file_stem):
                flags.append("generic_filename")
                inferred_title = extract_title_from_ocr(category_name, full_text)
                if inferred_title:
                    display_title = inferred_title
                    title_source = "ocr_header"

            if confidence_avg is not None and confidence_avg < 0.9:
                flags.append("low_confidence")

            record = ImageOcrRecord(
                category_index=category_index,
                category_name=category_name,
                file_stem=file_stem,
                display_title=display_title,
                title_source=title_source,
                image_path=str(image_path),
                image_name=image_path.name,
                full_text=full_text,
                confidence_avg=confidence_avg,
                line_count=len(lines),
                flags=flags,
            )
            records.append(record)
            items_payload.append(
                {
                    "rule_key": sanitize_identifier(display_title),
                    "image_name": image_path.name,
                    "file_stem": file_stem,
                    "display_title": display_title,
                    "title_source": title_source,
                    "ocr_confidence_avg": confidence_avg,
                    "flags": flags,
                    "ocr_lines": lines,
                    "ocr_full_text": full_text,
                    "image_path": str(image_path),
                }
            )

        per_category_payload.append(
            {
                "category_index": category_index,
                "category_name": category_name,
                "item_count": len(items_payload),
                "items": items_payload,
            }
        )

    catalog_path = output_dir / "catalog.json"
    catalog_path.write_text(
        json.dumps(
            {
                "source": "xiaohongshu_creator_center_manual",
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "category_count": len(per_category_payload),
                "item_count": len(records),
                "categories": per_category_payload,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    summary_path = output_dir / "summary.json"
    summary_path.write_text(
        json.dumps(
            [
                {
                    "category_index": record.category_index,
                    "category_name": record.category_name,
                    "display_title": record.display_title,
                    "title_source": record.title_source,
                    "image_name": record.image_name,
                    "ocr_confidence_avg": record.confidence_avg,
                    "flags": record.flags,
                    "image_path": record.image_path,
                }
                for record in records
            ],
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    workbook_path = SPREADSHEET_DIR / f"xiaohongshu_creator_center_manual_{timestamp}.xlsx"
    SPREADSHEET_DIR.mkdir(parents=True, exist_ok=True)
    build_workbook(records, workbook_path)

    print(
        json.dumps(
            {
                "catalog": str(catalog_path),
                "workbook": str(workbook_path),
                "item_count": len(records),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
